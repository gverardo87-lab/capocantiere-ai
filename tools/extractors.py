from __future__ import annotations

import os
import sys
# Import custom per aggiungere la root del progetto al path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import hashlib
import io
import json
import mimetypes
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict, Any

import pandas as pd
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader
from ollama import Client

from core.config import OLLAMA_MODEL


# --- Strutture Dati di Base ---
@dataclass
class ExtractedField:
    name: str
    value: Optional[str]
    confidence: str
    method: str


# --- Funzioni di Utilità ---
def file_sha256(data: bytes) -> str:
    h = hashlib.sha256()
    h.update(data)
    return h.hexdigest()

def sniff_mime(filename: str) -> str:
    m, _ = mimetypes.guess_type(filename)
    return m or "application/octet-stream"


# --- Motore di Estrazione Testo dai File ---
def extract_text_from_pdf(data: bytes) -> str:
    with io.BytesIO(data) as f:
        reader = PdfReader(f)
        return "\n".join(page.extract_text() or "" for page in reader.pages)

def extract_text_from_docx(data: bytes) -> str:
    with io.BytesIO(data) as f:
        doc = DocxDocument(f)
        return "\n".join(p.text for p in doc.paragraphs if p.text)

def extract_text_from_xlsx(data: bytes, max_cells: int = 2000) -> str:
    with io.BytesIO(data) as f:
        wb = load_workbook(f, data_only=True)
        text_chunks = []
        cell_count = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if value is not None:
                        text_chunks.append(str(value))
                        cell_count += 1
                        if cell_count >= max_cells:
                            return "\n".join(text_chunks)
        return "\n".join(text_chunks)


# --- FUNZIONE CORRETTA PER IL PARSING DEL CSV (VERSIONE MIGLIORATA) ---
def parse_timesheet_csv(data: bytes) -> Tuple[List[Dict[str, Any]], List[ExtractedField]]:
    """
    Legge un rapportino ore da un file CSV, valida e normalizza gli orari,
    e gestisce una colonna 'pausa' opzionale.
    """
    REQUIRED_COLUMNS = {"data", "commessa", "operaio", "orario_ingresso", "orario_uscita"}

    def _normalize_time_string(time_str: str, row_index: int) -> str:
        """Funzione helper per pulire e validare le stringhe di orario."""
        if not isinstance(time_str, str):
            raise ValueError(f"Riga {row_index + 2}: l'orario non è una stringa di testo valida.")
        
        # Sostituisce punti o altri separatori comuni con ':' e rimuove spazi
        cleaned_time = time_str.strip().replace('.', ':')
        
        # Prova a parsare l'orario per validarlo
        try:
            # Aggiunge i minuti se mancanti (es. "8" -> "8:00")
            if ':' not in cleaned_time:
                if len(cleaned_time) in [1, 2]: # Formato "8" o "14"
                     cleaned_time = f"{cleaned_time}:00"
            
            # Converte in formato HH:MM standard
            parsed_time = pd.to_datetime(cleaned_time, format='%H:%M').strftime('%H:%M')
            return parsed_time
        except ValueError:
            raise ValueError(
                f"Riga {row_index + 2}: formato ora non valido ('{time_str}'). "
                f"Usa un formato chiaro come 'HH:MM' (es. '08:00' o '17:30')."
            )

    with io.BytesIO(data) as f:
        df = pd.read_csv(f, sep=None, engine='python', dtype=str)

    df.columns = df.columns.str.lower().str.strip()
    
    actual_columns = set(df.columns)
    if not REQUIRED_COLUMNS.issubset(actual_columns):
        raise ValueError(f"Il file CSV non è un rapportino valido. Colonne obbligatorie mancanti: {REQUIRED_COLUMNS - actual_columns}")

    # --- NUOVA GESTIONE PAUSA ---
    # Se la colonna 'pausa' non esiste, la creiamo con il valore di default 1.0
    if 'pausa' not in df.columns:
        df['pausa'] = 1.0
    else:
        # Se esiste, la convertiamo in un numero, gestendo errori e valori mancanti
        df['pausa'] = pd.to_numeric(df['pausa'].str.replace(',', '.'), errors='coerce').fillna(1.0)

    # Standardizzazione delle altre colonne
    df['data'] = pd.to_datetime(df['data'], errors='coerce').dt.strftime('%Y-%m-%d')
    df = df.dropna(subset=['data', 'orario_ingresso', 'orario_uscita'])
    
    # Applichiamo la normalizzazione degli orari riga per riga per avere errori parlanti
    for index, row in df.iterrows():
        df.loc[index, 'orario_ingresso'] = _normalize_time_string(row['orario_ingresso'], index)
        df.loc[index, 'orario_uscita'] = _normalize_time_string(row['orario_uscita'], index)

    for col in ['reparto', 'descrizione']:
        if col not in df.columns:
            df[col] = ''
        df[col] = df[col].fillna('')

    summary_fields = []
    
    # Rinominiamo la colonna 'pausa' per coerenza con il database
    df.rename(columns={'pausa': 'durata_pausa_ore'}, inplace=True)
    
    final_columns = ['data', 'commessa', 'operaio', 'reparto', 'orario_ingresso', 'orario_uscita', 'descrizione', 'durata_pausa_ore']
    structured_rows = df[final_columns].to_dict('records')
    
    return structured_rows, summary_fields


# --- Funzione di Estrazione con AI (invariata) ---
def extract_fields_with_ai(text: str, kind: str) -> List[ExtractedField]:
    # ... (il resto di questa funzione rimane uguale)
    if not text or len(text.strip()) < 10:
        return []
    print(f"--- Avvio estrazione AI (Modello: {OLLAMA_MODEL}, Tipo Doc: {kind}) ---")
    client = Client()
    if kind == "FATTURA":
        prompt_fields = "'numero_fattura', 'data_documento', 'totale_imponibile', 'totale_iva', 'totale_fattura', 'partita_iva_fornitore'"
    else:
        prompt_fields = "'oggetto_documento', 'data_documento', 'nome_cliente', 'riferimento_interno'"
    prompt = f"""
    Sei un assistente AI specializzato nell'analisi di documenti aziendali.
    Dal testo fornito, estrai i seguenti campi: {prompt_fields}.
    Rispondi ESCLUSIVAMENTE con un oggetto JSON valido. Non aggiungere spiegazioni o testo introduttivo.
    Se un valore non viene trovato, assegna il valore null a quel campo.

    Testo da analizzare:
    ---
    {text[:4000]}
    ---
    """
    try:
        response = client.chat(
            model=OLLAMA_MODEL,
            messages=[{'role': 'user', 'content': prompt}],
            format="json",
            options={'temperature': 0.0}
        )
        content = response['message']['content']
        print(f"Risposta JSON grezza dal modello: {content}")
        data = json.loads(content)
        fields = []
        if isinstance(data, dict):
            for key, value in data.items():
                if value:
                    fields.append(ExtractedField(
                        name=key.strip().lower(),
                        value=str(value).strip(),
                        confidence="yellow",
                        method=OLLAMA_MODEL.split(":")[0]
                    ))
        print(f"--- Estrazione AI completata. Trovati {len(fields)} campi. ---")
        return fields
    except Exception as e:
        print(f"ERRORE: Impossibile completare l'estrazione AI: {e}")
        return []


# --- Logica di Classificazione (invariata) ---
def classify_kind(text: str) -> str:
    # ... (il resto di questa funzione rimane uguale)
    t = (text or "").lower()
    if any(k in t for k in ["fattura", "imponibile", "iva"]):
        return "FATTURA"
    if any(k in t for k in ["rapporto", "rapportino", "ore lavorate"]):
        return "RAPPORTO"
    if any(k in t for k in ["permesso di lavoro", "work permit"]):
        return "PERMESSO"
    return "ALTRO"


# --- Funzione Principale di Orchestrazione (invariata) ---
def read_text_and_kind(filename: str, data: bytes) -> Tuple[str, str]:
    """
    Orchestrates text extraction and document classification.
    Gives priority to structural checks (CSV headers) before content-based classification.
    """
    if filename.lower().endswith('.csv'):
        try:
            with io.BytesIO(data) as f:
                header = pd.read_csv(f, nrows=0, sep=None, engine='python').columns.str.lower().to_list()
            if {"data", "commessa", "operaio", "orario_ingresso", "orario_uscita"}.issubset(set(header)):
                return data.decode('utf-8', errors='ignore'), "RAPPORTO_CSV"
        except Exception as e:
            print(f"INFO: Could not parse CSV header for '{filename}'. Falling back to text extraction. Error: {e}")
            pass

    mime_type = sniff_mime(filename)
    text = ""
    if mime_type == "application/pdf":
        text = extract_text_from_pdf(data)
    elif "wordprocessingml" in mime_type or filename.endswith(".docx"):
        text = extract_text_from_docx(data)
    elif "spreadsheetml" in mime_type or filename.endswith(".xlsx"):
        text = extract_text_from_xlsx(data)
    else:
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            text = data.decode("latin-1", errors="ignore")

    kind = classify_kind(text)
    return text, kind